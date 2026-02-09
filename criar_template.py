import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from datetime import datetime

def criar_template_entrada_dados():
    """
    Cria um template Excel formatado para entrada de dados mensais de vagas
    """
    
    # Criar um novo workbook
    wb = Workbook()
    
    # ===========================================
    # ABA 1: INSTRUÇÕES
    # ===========================================
    ws_instrucoes = wb.active
    ws_instrucoes.title = "INSTRUÇÕES"
    
    # Estilos
    header_fill = PatternFill(start_color="76B82A", end_color="76B82A", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=14)
    instrucao_font = Font(size=11)
    
    instruções = [
        ["COMO USAR ESTE TEMPLATE"],
        [""],
        ["1. Vá até a aba 'DADOS_MENSAIS'"],
        ["2. Preencha as informações de cada MÊS que você deseja adicionar"],
        ["3. Preencha todas as células amarelas com os valores corretos"],
        ["4. NÃO altere os cabeçalhos (primeira linha)"],
        ["5. Salve o arquivo com um nome descritivo (ex: dados_marco_2026.xlsx)"],
        ["6. Execute o script Python para atualizar os gráficos"],
        [""],
        ["IMPORTANTE:"],
        ["- As células em AMARELO devem ser preenchidas"],
        ["- As células em CINZA não devem ser alteradas"],
        ["- Use apenas NÚMEROS nas células de dados (sem texto)"],
        ["- O campo 'Mês' deve estar no formato: 'Janeiro', 'Fevereiro', 'Março', etc."],
        [""],
        ["EXEMPLO DE PREENCHIMENTO:"],
        ["Mês: Março"],
        ["SP_Entradas: 150"],
        ["SP_Saidas: 180"],
        ["MS_Entradas: 60"],
        ["MS_Saidas: 45"],
    ]
    
    for i, linha in enumerate(instruções, start=1):
        ws_instrucoes.cell(row=i, column=1, value=linha[0])
        if i == 1:
            ws_instrucoes.cell(row=i, column=1).fill = header_fill
            ws_instrucoes.cell(row=i, column=1).font = header_font
        else:
            ws_instrucoes.cell(row=i, column=1).font = instrucao_font
    
    ws_instrucoes.column_dimensions['A'].width = 80
    
    # ===========================================
    # ABA 2: DADOS MENSAIS
    # ===========================================
    ws_dados = wb.create_sheet(title="DADOS_MENSAIS")
    
    # Cabeçalhos
    headers = ["Mês", "SP_Entradas", "SP_Saidas", "MS_Entradas", "MS_Saidas", "Observações"]
    
    # Estilos para cabeçalho
    header_fill_azul = PatternFill(start_color="30515F", end_color="30515F", fill_type="solid")
    header_font_white = Font(bold=True, color="FFFFFF", size=12)
    center_alignment = Alignment(horizontal="center", vertical="center")
    
    # Preencher cabeçalho
    for col_num, header in enumerate(headers, start=1):
        cell = ws_dados.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill_azul
        cell.font = header_font_white
        cell.alignment = center_alignment
    
    # Estilos para células de dados
    input_fill = PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid")  # Amarelo claro
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Adicionar linhas de exemplo (3 meses)
    meses_exemplo = [
        ["Março", "", "", "", "", ""],
        ["Abril", "", "", "", "", ""],
        ["Maio", "", "", "", "", ""],
    ]
    
    for row_num, dados_linha in enumerate(meses_exemplo, start=2):
        for col_num, valor in enumerate(dados_linha, start=1):
            cell = ws_dados.cell(row=row_num, column=col_num, value=valor)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Células de entrada (amarelo)
            if col_num in [1, 2, 3, 4, 5]:  # Mês, SP_Entradas, SP_Saidas, MS_Entradas, MS_Saidas
                cell.fill = input_fill
    
    # Ajustar largura das colunas
    ws_dados.column_dimensions['A'].width = 15
    ws_dados.column_dimensions['B'].width = 15
    ws_dados.column_dimensions['C'].width = 15
    ws_dados.column_dimensions['D'].width = 15
    ws_dados.column_dimensions['E'].width = 15
    ws_dados.column_dimensions['F'].width = 30
    
    # ===========================================
    # ABA 3: DADOS HISTÓRICOS (Referência)
    # ===========================================
    ws_historico = wb.create_sheet(title="DADOS_HISTÓRICOS")
    
    # Dados históricos atuais do sistema
    dados_historicos = {
        'Mês': ['Dezembro', 'Janeiro', 'Fevereiro'],
        'SP_Entradas': [45, 164, 54],
        'SP_Saidas': [73, 163, 199],
        'MS_Entradas': [20, 25, 14],
        'MS_Saidas': [40, 40, 37],
        'SP_Saldo_Periodo': [], 
        'MS_Saldo_Periodo': []
    }
    
    # Calcular saldos
    for i in range(len(dados_historicos['Mês'])):
        dados_historicos['SP_Saldo_Periodo'].append(
            dados_historicos['SP_Entradas'][i] - dados_historicos['SP_Saidas'][i]
        )
        dados_historicos['MS_Saldo_Periodo'].append(
            dados_historicos['MS_Entradas'][i] - dados_historicos['MS_Saidas'][i]
        )
    
    df_historico = pd.DataFrame(dados_historicos)
    
    # Escrever DataFrame na planilha
    for col_num, col_name in enumerate(df_historico.columns, start=1):
        cell = ws_historico.cell(row=1, column=col_num, value=col_name)
        cell.fill = header_fill_azul
        cell.font = header_font_white
        cell.alignment = center_alignment
    
    # Preencher dados
    gray_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
    
    for row_num, row_data in enumerate(df_historico.values, start=2):
        for col_num, value in enumerate(row_data, start=1):
            cell = ws_historico.cell(row=row_num, column=col_num, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.fill = gray_fill  # Cinza = somente leitura
    
    # Ajustar largura
    for col in ws_historico.columns:
        ws_historico.column_dimensions[col[0].column_letter].width = 18
    
    # Salvar arquivo
    timestamp = datetime.now().strftime('%Y%m%d')
    nome_arquivo = f'TEMPLATE_ENTRADA_DADOS_{timestamp}.xlsx'
    wb.save(nome_arquivo)
    
    print(f"[OK] Template criado com sucesso: {nome_arquivo}")
    print(f"[INFO] O arquivo possui 3 abas:")
    print(f"   1. INSTRUCOES - Como usar o template")
    print(f"   2. DADOS_MENSAIS - Preencha aqui os novos meses")
    print(f"   3. DADOS_HISTORICOS - Referencia dos dados atuais")
    print(f"\n[DICA] Envie este arquivo para a Denise preencher!")
    
    return nome_arquivo

if __name__ == "__main__":
    criar_template_entrada_dados()
